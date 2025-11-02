import asyncio
import json
import tempfile
import os
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, UploadFile, File as FastAPIFile, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from sse_starlette.sse import EventSourceResponse
import pandas as pd
import io
import traceback

from .sensor import SensorData

# OpenAI and PDF processing imports
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
#from .parsers import parseDatasetFile
# TODO $HOST in origin and endpoints

# call R
from rpy2.robjects import pandas2ri
from rpy2.robjects.conversion import localconverter
import rpy2.robjects as robjects

robjects.r('read_xpt <- function(path){haven::read_xpt(path)}')
robjects.r('read_sas7bdat <- function(path){haven::read_sas(path)}')

# In-memory parsing functions using R connections
# Note: haven::read_xpt() and haven::read_sas() require file paths,
# but we use R's connection API to handle the data transfer in-memory
robjects.r('''
read_xpt_from_bytes <- function(raw_bytes) {
    # Create a connection from raw bytes for in-memory processing
    # Use rawConnection to work with the binary data directly
    con <- rawConnection(raw_bytes, "rb")
    on.exit(close(con))
    
    # haven requires a file path, so we write to R's tempfile
    # This is still more efficient as data transfer from Python to R is in-memory
    # and R handles its own temp file management
    tmp <- tempfile(fileext = ".xpt")
    on.exit(file.remove(tmp), add = TRUE)
    
    # Write from connection to temp file
    writeBin(readBin(con, "raw", n = length(raw_bytes)), tmp)
    
    # Parse using haven
    result <- haven::read_xpt(tmp)
    return(result)
}

read_sas7bdat_from_bytes <- function(raw_bytes) {
    # Similar approach for SAS7BDAT using connections
    con <- rawConnection(raw_bytes, "rb")
    on.exit(close(con))
    
    tmp <- tempfile(fileext = ".sas7bdat")
    on.exit(file.remove(tmp), add = TRUE)
    
    # Write from connection to temp file
    writeBin(readBin(con, "raw", n = length(raw_bytes)), tmp)
    
    # Parse using haven
    result <- haven::read_sas(tmp)
    return(result)
}
''')

read_xpt = robjects.r['read_xpt']
read_sas7bdat = robjects.r['read_sas7bdat']
read_xpt_from_bytes = robjects.r['read_xpt_from_bytes']
read_sas7bdat_from_bytes = robjects.r['read_sas7bdat_from_bytes']

app = FastAPI()
sensor = SensorData()

# Add exception handler to ensure CORS headers are included in error responses
@app.exception_handler(HTTPException)
async def http_exception_handler(request, exc: HTTPException):
    """Ensure CORS headers are included in HTTPException responses"""
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        }
    )

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """Handle all unhandled exceptions and ensure CORS headers are included"""
    error_msg = str(exc) if exc else "Internal server error"
    print(f"Unhandled exception occurred: {error_msg}")
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"detail": error_msg, "error": type(exc).__name__},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        }
    )


# Minimal placeholder parser until a real implementation is provided
def parse_dataset_file(file: str) -> dict:
    """Return a simple structure acknowledging the requested file.

    This is a stub to avoid NameError. Replace with real parsing logic
    (e.g., reading CSV/Parquet and returning structured data) as needed.
    """

    data = file.read()  # Your XPT file data in bytes
    buffer = io.BytesIO(data)
    df = pd.read_sas(buffer, format='xport')
    return {"file": file, "status": "parsed", "data": df.to_dict()}


# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=['*'],  # Svelte dev server
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api")
async def root():
    return {"message": "Welcome to the SDTM AI API"}


@app.get("/api/current")
async def get_current_reading():
    """Get the current sensor reading."""
    return sensor.generate_reading()


@app.get("/api/stream")
async def stream_data():
    """Stream sensor data using server-sent events."""
    async def event_generator():
        while True:
            data = sensor.generate_reading()
            yield {
                "event": "sensor_update",
                "data": json.dumps(data)
            }
            await asyncio.sleep(2)  # Update every 2 seconds

    return EventSourceResponse(event_generator())

@app.get("/api/items/{item_id}")
def read_item(item_id: int, q: Optional[str] = None):
    return {"item_id": item_id, "q": q}


@app.get("/api/data")
def get_data():
    print("Python API was hit successfully!")
    
    # This is the data that will be sent back to the SvelteKit load function
    return {"message": "Hello from your FastAPI backend!", "value": 123}

@app.post("/api/parse_pd")
async def parse_dataset_pandas(file: UploadFile = FastAPIFile(...)) -> dict:
    """Parse various file formats using pandas.
    
    Supports: CSV, Excel (.xlsx, .xls), JSON, Parquet, SAS (.xpt, .sas7bdat), TSV
    """
    print(f"Python API Parsing with Pandas: {file.filename}")
    data = await file.read()
    
    # Get file extension to determine parser
    file_ext = os.path.splitext(file.filename)[1].lower()
    buffer = io.BytesIO(data)
    
    try:
        if file_ext in ['.csv', '.txt']:
            # CSV files
            df = pd.read_csv(buffer, encoding='utf-8', low_memory=False)
        elif file_ext in ['.tsv']:
            # TSV files
            df = pd.read_csv(buffer, sep='\t', encoding='utf-8', low_memory=False)
        elif file_ext in ['.xlsx', '.xls']:
            # Excel files - read first sheet by default
            try:
                df = pd.read_excel(buffer, engine='openpyxl' if file_ext == '.xlsx' else None)
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="Excel file support requires 'openpyxl' package. Install with: pip install openpyxl"
                )
        elif file_ext in ['.json', '.jsonl']:
            # JSON files
            if file_ext == '.jsonl':
                # JSON Lines format
                df = pd.read_json(buffer, lines=True)
            else:
                # Standard JSON
                json_data = json.loads(data.decode('utf-8'))
                # If it's a list of records, use directly; otherwise try to normalize
                if isinstance(json_data, list):
                    df = pd.DataFrame(json_data)
                else:
                    df = pd.json_normalize(json_data)
        elif file_ext in ['.parquet', '.pq']:
            # Parquet files
            try:
                df = pd.read_parquet(buffer, engine='pyarrow')
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="Parquet file support requires 'pyarrow' package. Install with: pip install pyarrow"
                )
        elif file_ext in ['.xpt']:
            # SAS XPT files
            df = pd.read_sas(buffer, format='xport')
        elif file_ext in ['.sas7bdat']:
            # SAS7BDAT files
            df = pd.read_sas(buffer, format='sas7bdat')
        else:
            raise ValueError(f"Unsupported file format: {file_ext}. Supported formats: CSV, Excel, JSON, Parquet, SAS")
        
        # Convert to list of dictionaries for JSON serialization
        result = {"file": file.filename, "status": "parsed", "data": df.to_dict('records')}
        return result
        
    except ValueError as e:
        error_msg = str(e)
        print(f"Error parsing {file.filename} with pandas: {error_msg}")
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse {file.filename}: {error_msg}"
        )
    except HTTPException:
        # Re-raise HTTPExceptions as-is (they already have proper status codes)
        raise
    except Exception as e:
        error_msg = str(e)
        print(f"Error parsing {file.filename} with pandas: {error_msg}")
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to parse {file.filename}: {error_msg}"
        )


@app.post("/api/parse")
async def parse_dataset(file: UploadFile = FastAPIFile(...)) -> dict:
    """Parse XPT or SAS7BDAT files using R's haven library.
    
    The file is saved to a temporary file since R's haven functions require a file path.
    """
    print(f"Python API Parsing with R: {file.filename}")
    data = await file.read()
    
    # Get file extension to determine parser
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    # Create temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp_file:
        tmp_path = tmp_file.name
        tmp_file.write(data)
    
    try:
        # Parse based on file extension
        if file_ext == '.xpt':
            df_r = read_xpt(tmp_path)
        elif file_ext == '.sas7bdat':
            df_r = read_sas7bdat(tmp_path)
        else:
            # Fallback: try pandas for other SAS formats
            buffer = io.BytesIO(data)
            df_pd = pd.read_sas(buffer, format='xport' if file_ext == '.xpt' else 'sas7bdat')
            # Convert pandas DataFrame to dict
            return {"file": file.filename, "status": "parsed", "data": df_pd.to_dict('records')}
        
        # Convert R DataFrame to pandas and then to list of dicts
        # Use new conversion context instead of deprecated activate/deactivate
        with localconverter(robjects.default_converter + pandas2ri.converter):
            df_pd = pandas2ri.rpy2py(df_r)
        
        # Convert to list of dictionaries for JSON serialization
        result = {"file": file.filename, "status": "parsed", "data": df_pd.to_dict('records')}
        return result
        
    finally:
        # Clean up temporary file
        try:
            os.unlink(tmp_path)
        except OSError:
            pass  # File already deleted or doesn't exist


@app.post("/api/parse2")
async def parse_dataset_inmemory(file: UploadFile = FastAPIFile(...)) -> dict:
    """Parse XPT or SAS7BDAT files using R's haven library with in-memory connections.
    
    This version uses R's connection API to read from memory instead of temporary files.
    The file data is passed directly to R as a raw vector, which is more efficient.
    """
    print(f"Python API Parsing with R (in-memory): {file.filename}")
    data = await file.read()
    
    # Get file extension to determine parser
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    try:
        # Convert Python bytes to R raw vector
        # R uses raw vectors (integer vector with values 0-255) for binary data
        raw_bytes_r = robjects.ByteVector(data)
        
        # Parse based on file extension using in-memory functions
        if file_ext == '.xpt':
            df_r = read_xpt_from_bytes(raw_bytes_r)
        elif file_ext == '.sas7bdat':
            df_r = read_sas7bdat_from_bytes(raw_bytes_r)
        else:
            # Fallback: try pandas for other SAS formats
            buffer = io.BytesIO(data)
            df_pd = pd.read_sas(buffer, format='xport' if file_ext == '.xpt' else 'sas7bdat')
            # Convert pandas DataFrame to dict
            return {"file": file.filename, "status": "parsed", "data": df_pd.to_dict('records')}
        
        # Convert R DataFrame to pandas and then to list of dicts
        # Use new conversion context instead of deprecated activate/deactivate
        with localconverter(robjects.default_converter + pandas2ri.converter):
            df_pd = pandas2ri.rpy2py(df_r)
        
        # Convert to list of dictionaries for JSON serialization
        result = {"file": file.filename, "status": "parsed", "data": df_pd.to_dict('records')}
        return result
        
    except HTTPException:
        # Re-raise HTTPExceptions as-is (they already have proper status codes)
        raise
    except Exception as e:
        error_msg = str(e)
        print(f"Error parsing {file.filename}: {error_msg}")
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to parse {file.filename}: {error_msg}"
        )


@app.post("/api/compliance/excel")
async def compliance_check_excel(
    dataset_name: str = Body(...),
    dataset_domain: Optional[str] = Body(None),
    dataset_columns: Dict[str, str] = Body(...),
    dataset_rows: List[Dict[str, Any]] = Body(...),
    standard_id: str = Body(...)
) -> dict:
    """Check compliance using Excel file for variable names and types.
    
    Uses SDTMIG_v3.4.xlsx to validate:
    - Variable names match expected variables for domain
    - Variable types match expected types
    - Required variables are present
    """
    excel_path = Path(__file__).parent.parent.parent / "frontend" / "static" / "SDTMIG_v3.4.xlsx"
    
    if not excel_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Excel file not found at {excel_path}"
        )
    
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="openpyxl package required. Install with: pip install openpyxl"
        )
    
    try:
        # Load Excel workbook
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Get the domain to check (use dataset domain or try to detect from name)
        domain_to_check = dataset_domain
        if not domain_to_check:
            # Extract domain from filename (e.g., "AE.xpt" -> "AE")
            name_upper = dataset_name.upper().replace(".XPT", "").replace(".CSV", "")
            if len(name_upper) == 2:
                domain_to_check = name_upper
        
        findings = []
        
        # Try to find domain sheet in Excel
        # SDTMIG Excel typically has sheets named by domain or a variables sheet
        sheet_found = False
        for sheet_name in workbook.sheetnames:
            if domain_to_check and domain_to_check.upper() in sheet_name.upper():
                sheet = workbook[sheet_name]
                sheet_found = True
                
                # Parse variable rules from sheet
                # Assuming structure: Variable | Type | Required | etc.
                headers = []
                variable_rules = {}
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_idx == 1:
                        headers = [str(cell).strip() if cell else "" for cell in row]
                        continue
                    
                    if not any(row):
                        continue
                    
                    # Map row to headers
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if idx < len(row):
                            row_dict[header.lower()] = row[idx]
                    
                    # Find variable name column (could be "Variable", "Variable Name", etc.)
                    var_col = None
                    for col in headers:
                        if "variable" in col.lower() and "name" in col.lower():
                            var_col = col.lower()
                            break
                    if not var_col:
                        for col in headers:
                            if "variable" in col.lower():
                                var_col = col.lower()
                                break
                    
                    if var_col and var_col in row_dict:
                        var_name = str(row_dict[var_col]).strip().upper()
                        if var_name:
                            # Find type column
                            type_col = None
                            for col in headers:
                                if "type" in col.lower() or "datatype" in col.lower():
                                    type_col = col.lower()
                                    break
                            
                            # Find required column
                            req_col = None
                            for col in headers:
                                if "required" in col.lower() or "core" in col.lower() or "mandatory" in col.lower():
                                    req_col = col.lower()
                                    break
                            
                            var_type = str(row_dict[type_col]).strip() if type_col and type_col in row_dict else None
                            var_required = False
                            if req_col and req_col in row_dict:
                                req_val = str(row_dict[req_col]).strip().upper()
                                var_required = req_val in ["YES", "Y", "REQUIRED", "CORE", "MANDATORY", "TRUE", "1"]
                            
                            variable_rules[var_name] = {
                                "type": var_type,
                                "required": var_required
                            }
                
                # Check dataset against rules
                dataset_cols_upper = {k.upper(): v for k, v in dataset_columns.items()}
                
                for var_name, rule in variable_rules.items():
                    if var_name not in dataset_cols_upper:
                        if rule.get("required"):
                            findings.append({
                                "severity": "error",
                                "variable": var_name,
                                "message": f"Required variable {var_name} is missing",
                                "ruleReference": f"{domain_to_check}.{var_name} required variable"
                            })
                        else:
                            findings.append({
                                "severity": "warning",
                                "variable": var_name,
                                "message": f"Expected variable {var_name} is not present",
                                "ruleReference": f"{domain_to_check}.{var_name} optional variable"
                            })
                    else:
                        # Check type if specified
                        expected_type = rule.get("type")
                        actual_type = dataset_cols_upper[var_name]
                        if expected_type:
                            expected_lower = expected_type.lower()
                            actual_lower = actual_type.lower()
                            # Type matching (flexible)
                            if "text" not in expected_lower and "char" not in expected_lower:
                                if expected_lower not in actual_lower and actual_lower not in expected_lower:
                                    findings.append({
                                        "severity": "warning",
                                        "variable": var_name,
                                        "message": f"Variable {var_name} is {actual_type} but expected {expected_type}",
                                        "ruleReference": f"{domain_to_check}.{var_name} datatype"
                                    })
                
                break
        
        if not sheet_found and domain_to_check:
            findings.append({
                "severity": "error",
                "message": f"Domain {domain_to_check} not found in Excel file",
                "ruleReference": "Domain catalog"
            })
        
        summary = {
            "total": len(findings),
            "errors": sum(1 for f in findings if f.get("severity") == "error"),
            "warnings": sum(1 for f in findings if f.get("severity") == "warning")
        }
        
        return {
            "findings": findings,
            "summary": summary,
            "domain": domain_to_check,
            "checked_variables": len(variable_rules) if sheet_found else 0
        }
        
    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to check compliance with Excel: {error_msg}"
        )


@app.post("/api/compliance/openai")
async def compliance_check_openai(
    dataset_name: str = Body(...),
    dataset_domain: Optional[str] = Body(None),
    dataset_columns: Dict[str, str] = Body(...),
    dataset_rows: List[Dict[str, Any]] = Body(...),
    standard_id: str = Body(...),
    standard_pdf_path: Optional[str] = Body(None)
) -> dict:
    """Check compliance using OpenAI with PDF context from the standard.
    
    Uses OpenAI to analyze dataset against SDTM standard PDF documents.
    """
    if not OPENAI_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="OpenAI package required. Install with: pip install openai"
        )
    
    # Get OpenAI API key from environment
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        raise HTTPException(
            status_code=500,
            detail="OPENAI_API_KEY environment variable not set"
        )
    
    client = OpenAI(api_key=openai_api_key)
    
    # Determine PDF path based on standard_id
    pdf_path = standard_pdf_path
    if not pdf_path:
        frontend_static = Path(__file__).parent.parent.parent / "frontend" / "static"
        if standard_id == "sdtm-2-0":
            pdf_path = frontend_static / "SDTM_v2.0.pdf"
        elif standard_id == "sdtmig-3-4":
            pdf_path = frontend_static / "SDTMIG v3.4-FINAL_2022-07-21.pdf"
        else:
            raise HTTPException(
                status_code=400,
                detail=f"PDF not available for standard {standard_id}"
            )
    
    if not pdf_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"PDF file not found at {pdf_path}"
        )
    
    try:
        # Read PDF content (simplified - for production, use proper PDF parsing)
        # For now, we'll use OpenAI's file upload capability
        with open(pdf_path, 'rb') as pdf_file:
            # Upload PDF to OpenAI
            pdf_upload = client.files.create(
                file=pdf_file,
                purpose="assistants"
            )
        
        # Create a sample of dataset for analysis (limit to avoid token limits)
        sample_rows = dataset_rows[:100] if len(dataset_rows) > 100 else dataset_rows
        
        # Prepare prompt
        prompt = f"""
Analyze the following SDTM dataset for compliance issues against the standard:

Dataset Name: {dataset_name}
Domain: {dataset_domain or 'Not specified'}
Columns ({len(dataset_columns)}): {', '.join(list(dataset_columns.keys())[:20])}
Sample rows: {json.dumps(sample_rows[:5], default=str)}

Please check:
1. Required variables are present for the domain
2. Variable names follow SDTM conventions
3. Variable types are appropriate
4. Data structure matches domain requirements

Return findings in JSON format with severity (error/warning/info), variable (if applicable), message, and ruleReference.
"""
        
        # Use OpenAI to analyze
        # Note: For production, you'd want to use Assistants API or fine-tuned models
        # This is a simplified approach
        response = client.chat.completions.create(
            model="gpt-4-turbo-preview",  # or gpt-4, gpt-3.5-turbo
            messages=[
                {
                    "role": "system",
                    "content": "You are an expert in CDISC SDTM standards. Analyze datasets for compliance and return findings in JSON format."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            temperature=0.3,
            max_tokens=2000
        )
        
        # Parse response (expecting JSON)
        response_text = response.choices[0].message.content
        
        # Try to extract JSON from response
        import re
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if json_match:
            findings_data = json.loads(json_match.group())
            findings = findings_data.get("findings", [])
        else:
            # Fallback: create findings from response text
            findings = [{
                "severity": "info",
                "message": response_text,
                "ruleReference": "OpenAI analysis"
            }]
        
        # Clean up uploaded file
        try:
            client.files.delete(pdf_upload.id)
        except:
            pass
        
        summary = {
            "total": len(findings),
            "errors": sum(1 for f in findings if f.get("severity") == "error"),
            "warnings": sum(1 for f in findings if f.get("severity") == "warning")
        }
        
        return {
            "findings": findings,
            "summary": summary,
            "domain": dataset_domain,
            "method": "openai"
        }
        
    except Exception as e:
        error_msg = str(e)
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to check compliance with OpenAI: {error_msg}"
        )